import openpyxl

w7a_pre_questions = [
    {
        "q": "พอร์ต 22 บนระบบ Linux Server มักถูกสงวนไว้สำหรับการทำงานของบริการข้อใด?",
        "options": [
            "HTTP Web Server (หน้าเว็บไม่เข้ารหัส)",
            "SSH (Secure Shell) สำหรับควบคุมระยะไกล",
            "MariaDB / MySQL Database Server",
            "DNS Name Resolution สำหรับแปลงชื่อโดเมน"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ข้อใดอธิบายความแตกต่างระหว่าง TCP และ UDP ได้ถูกต้องที่สุด?",
        "options": [
            "UDP มีระบบตรวจสอบความครบถ้วนของข้อมูลก่อนส่ง แต่ TCP ไม่มี",
            "TCP ทำงานได้รวดเร็วกว่า UDP เนื่องจากไม่ต้องมีการตรวจสอบสัญญาณตอบกลับ",
            "TCP ต้องทำการ Handshake (จับมือตกลงเชื่อมต่อ) ก่อนส่งข้อมูลจริง แต่ UDP ไม่ต้อง",
            "UDP เหมาะสำหรับบริการแชร์ไฟล์ที่ห้ามมีข้อมูลสูญหายเด็ดขาด"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "พอร์ตมาตรฐานสำหรับการให้บริการเว็บไซต์แบบธรรมดาที่ไม่มีการเข้ารหัสความปลอดภัย (HTTP) คือพอร์ตใด?",
        "options": [
            "พอร์ต 22",
            "พอร์ต 80",
            "พอร์ต 443",
            "พอร์ต 3306"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "พอร์ตมาตรฐานสำหรับการให้บริการเว็บไซต์แบบปลอดภัยที่มีการเข้ารหัสความปลอดภัย (HTTPS) คือพอร์ตใด?",
        "options": [
            "พอร์ต 22",
            "พอร์ต 80",
            "พอร์ต 443",
            "พอร์ต 3306"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "พอร์ตหมายเลข 3306 มักเกี่ยวข้องกับการให้บริการของซอฟต์แวร์ประเภทใดบนเซิร์ฟเวอร์?",
        "options": [
            "Nginx Web Server",
            "OpenSSH Server",
            "MariaDB / MySQL Database Server",
            "Samba File Sharing Server"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "หากต้องการตรวจสอบว่าระบบ Linux ของเราเปิดบริการพอร์ตใดค้างไว้เพื่อรอรับสายการเชื่อมต่อ (Listen) อยู่ ควรใช้คำสั่งใด?",
        "options": [
            "ip address show",
            "ss -tulpn",
            "ping -c 4 localhost",
            "curl -I localhost"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "การกำหนด Socket Binding ของโปรแกรมบริการไปที่ที่อยู่ 127.0.0.1 มีความหมายอย่างไร?",
        "options": [
            "อนุญาตการเชื่อมต่อจากอุปกรณ์ภายนอกระบบทั้งหมดโดยไม่ต้องยืนยันตัวตน",
            "อนุญาตให้เชื่อมต่อได้เฉพาะจากโปรแกรมที่รันอยู่ภายในเซิร์ฟเวอร์เครื่องเดียวกันเท่านั้น",
            "เป็นการปิดพอร์ตบริการนั้นๆ ชั่วคราวบนระบบปฏิบัติการ",
            "อนุญาตให้เชื่อมต่อเข้ามาเฉพาะจากอุปกรณ์เครือข่ายภายในวงแลนเดียวกัน"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ที่อยู่ไอพี (IP Address) ของอุปกรณ์เราเตอร์ที่ทำหน้าที่เป็นประตูทางออกนอกเครือข่ายอินเทอร์เน็ต เรียกว่าอะไร?",
        "options": [
            "Subnet Mask",
            "Default Gateway",
            "DNS Server",
            "Loopback Address"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "ซอฟต์แวร์เครื่องมือ Nmap (Network Mapper) ถูกนำมาใช้ประโยชน์ในข้อใดมากที่สุด?",
        "options": [
            "ตรวจสอบและวิเคราะห์การใช้งานพื้นที่บนฮาร์ดดิสก์",
            "ใช้สแกนเครือข่ายเพื่อค้นหาพอร์ตและบริการที่เปิดอยู่รวมถึงช่องโหว่",
            "ใช้จัดการสิทธิ์ผู้ใช้อันดับสูงภายในระบบปฏิบัติการ",
            "เปิดใช้งานหน้าต่าง Firewall ของ Ubuntu"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "การ Bind บริการเครือข่ายไว้ที่ไอพี 0.0.0.0 มีความหมายว่าอย่างไร?",
        "options": [
            "ยินดีรับสายและเชื่อมต่อจากทุก IP Interface ที่ชี้มาหาเครื่องเซิร์ฟเวอร์นี้",
            "จำกัดสิทธิ์เฉพาะเครื่องแม่ข่ายกลางเท่านั้นที่มีสิทธิ์สั่งการ",
            "ไอพีเครือข่ายจะทำงานเฉพาะเมื่อเซิร์ฟเวอร์เปิดหน้าเว็บพอร์ต 80 เท่านั้น",
            "ระบบจะทำลายแพ็กเก็ตทราฟฟิกข้อมูลที่ส่งมาผิดพลาดทิ้งทันที"
        ],
        "correct": 1,
        "time": 30
    }
]

w7a_post_questions = [
    {
        "q": "หากผลลัพธ์จากคำสั่ง ss -tulpn ระบุพอร์ต 3306 ผูกอยู่กับ localhost (127.0.0.1:3306) ผลลัพธ์ในแง่ความปลอดภัยเป็นอย่างไร?",
        "options": [
            "เครื่องอื่นในระบบยังเชื่อมเข้าใช้งานฐานข้อมูลได้ตามปกติ",
            "ปลอดภัยมาก เพราะบุคคลภายนอกไม่สามารถเข้าถึงพอร์ต 3306 ได้โดยตรงแม้ไม่มีไฟร์วอลล์",
            "ทำให้โปรแกรม PHP/NodeJS ในเครื่องเดียวกันไม่สามารถดึงข้อมูลได้",
            "พอร์ตจะถูกบล็อกและปิดบริการไปโดยอัตโนมัติ"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "หากใช้ Nmap สแกนพอร์ตเป้าหมายแล้วได้รับสถานะพอร์ตระบุเป็น 'filtered' หมายความว่าอย่างไร?",
        "options": [
            "พอร์ตนี้ถูกปิดและไม่มีบริการใดๆ ทำงานอยู่เบื้องหลัง",
            "พอร์ตเปิดทำงานอยู่และไม่มีการคัดกรองความปลอดภัยใดๆ",
            "มีไฟร์วอลล์คอยบล็อกกั้นแพ็กเก็ตสแกน ทำให้ตรวจสถานะจริงไม่ได้",
            "บริการในพอร์ตนี้มีการเข้ารหัสความปลอดภัยระดับโปรโตคอล TLS"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "คำสั่งสแกน Nmap ในข้อใดใช้สแกนพอร์ตพร้อมตรวจสอบเวอร์ชันของซอฟต์แวร์ที่รันอยู่เบื้องหลัง (Service Version Detection)?",
        "options": [
            "nmap -sV 192.168.1.100",
            "nmap -p 1-100 192.168.1.100",
            "nmap localhost",
            "nmap -v 192.168.1.100"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "คำสั่งใดใช้ตรวจสอบรายละเอียด IP Address ของอุปกรณ์เน็ตเวิร์กการ์ด (Network Interface) ในเซิร์ฟเวอร์ Ubuntu?",
        "options": [
            "ip a",
            "hostname -I",
            "ip route",
            "ss -tulpn"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "คำสั่ง 'telnet 192.168.1.100 80' มีประโยชน์อย่างไรสำหรับ System Administrator?",
        "options": [
            "ใช้ควบคุม CLI ของเครื่อง 192.168.1.100 ระยะไกล",
            "ใช้ตรวจสอบว่าพอร์ต 80 ของไอพีเป้าหมายเปิดรับสายหรือไม่",
            "ใช้วัดความเร็วอินเทอร์เน็ตในการส่งข้อมูลไปยัง Nginx",
            "ใช้สั่งปิดการทำงานพอร์ตเว็บเซิร์ฟเวอร์แบบฉุกเฉิน"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ในการทำงานของโปรโตคอล TCP ข้อมูลเริ่มแรกที่จะทำการขอเปิดการเชื่อมต่อ (Handshake 1) จะใช้แพ็กเก็ตชนิดใด?",
        "options": [
            "ACK (Acknowledge)",
            "SYN (Synchronize)",
            "SYN-ACK (Synchronize-Acknowledge)",
            "FIN (Finish)"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "พอร์ตหมายเลข 445 บนระบบ Linux Server มักถูกสงวนไว้สำหรับการทำงานของบริการข้อใด?",
        "options": [
            "SSH (Secure Shell) สำหรับควบคุมระยะไกล",
            "Nginx HTTP Web Server",
            "Samba Server สำหรับแชร์ไฟล์ข้ามระบบปฏิบัติการ",
            "MariaDB Database Server"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "หากต้องการตรวจสอบ Default Gateway บนตารางเส้นทางข้ามเครือข่ายของเซิร์ฟเวอร์ลินุกซ์ด้วย CLI ควรใช้คำสั่งใด?",
        "options": [
            "ip route",
            "ip a show",
            "ping localhost",
            "ss -tulpn"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "ข้อใดเป็นคุณสมบัติเด่นที่ถูกต้องของการรับส่งข้อมูลแบบ UDP (User Datagram Protocol)?",
        "options": [
            "รับประกันการส่งข้อมูลถึงผู้รับครบถ้วนโดยไม่มีข้อมูลสูญหาย",
            "เน้นความเร็วสูงสุดโดยไม่มีขั้นตอนต่อสาย (Handshake) ก่อนส่งข้อมูล",
            "บังคับเชื่อมต่อสัญญาณระดับฮาร์ดแวร์ก่อนส่งข้อมูลเสมอ",
            "มีการจัดเรียงลำดับชุดข้อมูลให้ถูกต้องก่อนแสดงผลปลายทาง"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ระบบ Firewall ส่วนใหญ่ทำงานในการกลั่นกรองอนุญาตหรือปิดกั้นข้อมูล โดยอิงจากข้อมูลหลักในข้อใด (OSI Layer 3-4)?",
        "options": [
            "MAC Address ของตัวเครื่องและสายแลน",
            "หมายเลขไอพีเครื่องต้นทาง/ปลายทาง และหมายเลขพอร์ตบริการ",
            "HTTP Request Method และความยาวของไฟล์",
            "ชื่อบัญชีของระบบปฏิบัติการและระดับสิทธิ์ไฟล์"
        ],
        "correct": 2,
        "time": 30
    }
]

def generate_xlsx(questions, output_path):
    try:
        wb = openpyxl.load_workbook('./documents/QuizizzSampleSpreadsheetUpdated_v2.xlsx')
        ws = wb.active
        
        # Clear rows
        for r in range(3, 100):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None
                
        # Populate
        for idx, q_data in enumerate(questions):
            row = idx + 3
            ws.cell(row=row, column=1).value = q_data["q"]
            ws.cell(row=row, column=2).value = "Multiple Choice"
            ws.cell(row=row, column=3).value = q_data["options"][0]
            ws.cell(row=row, column=4).value = q_data["options"][1]
            ws.cell(row=row, column=5).value = q_data["options"][2]
            ws.cell(row=row, column=6).value = q_data["options"][3]
            ws.cell(row=row, column=7).value = None
            ws.cell(row=row, column=8).value = q_data["correct"]
            ws.cell(row=row, column=9).value = q_data["time"]
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None
            
        wb.save(output_path)
        print(f"Successfully generated {output_path} with {len(questions)} questions")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    generate_xlsx(w7a_pre_questions, './public/data/week-7a_pre_wayground.xlsx')
    generate_xlsx(w7a_post_questions, './public/data/week-7a_post_wayground.xlsx')
