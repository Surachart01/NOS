import openpyxl

w6a_questions = [
    {
        "q": "การเข้าสู่ระบบปฏิบัติการ Linux Server ด้วยบัญชี 'root' ในการทำงานปกติขัดกับหลักการความปลอดภัยในข้อใด?",
        "options": [
            "Principle of Least Privilege (หลักการมอบสิทธิ์ขั้นต่ำที่จำเป็น)",
            "Multi-factor Authentication (การยืนยันตัวตนหลายปัจจัย)",
            "Data Encapsulation (การห่อหุ้มข้อมูล)",
            "System Performance Optimization (การปรับแต่งประสิทธิภาพระบบ)"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "หากต้องการสร้างผู้ใช้งานใหม่ชื่อ 'student01' บน Ubuntu Server พร้อมสร้าง Home Directory ให้โดยอัตโนมัติ ควรใช้คำสั่งใด?",
        "options": [
            "sudo useradd student01",
            "sudo adduser student01",
            "sudo newuser student01",
            "sudo makeuser student01"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "คำสั่งในข้อใดใช้เพื่อจัดกลุ่มและนำผู้ใช้งาน 'student01' เข้าสู่กลุ่มพิเศษ 'sudo' เพื่อให้สามารถยกระดับสิทธิ์ได้?",
        "options": [
            "sudo usermod -aG sudo student01",
            "sudo useradd sudo student01",
            "sudo groupadd student01 sudo",
            "sudo chmod +x student01 sudo"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "ในระบบไฟล์ของ Linux ข้อความสิทธิ์ '-rw-r--r--' มีความหมายอย่างไร?",
        "options": [
            "เจ้าของและกลุ่มอ่านเขียนรันได้ แต่คนอื่นอ่านได้อย่างเดียว",
            "ทุกคนในระบบสามารถเปิดอ่าน แก้ไข และสั่งรันไฟล์นี้ได้อิสระ",
            "เจ้าของอ่านเขียนได้ ส่วนกลุ่มและผู้อื่นอ่านได้อย่างเดียวโดยห้ามแก้ไข",
            "ไฟล์นี้ถูกล็อกสิทธิ์ไม่ให้ผู้ใดแก้ไขหรือลบได้ถาวร"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "หากต้องการกำหนดสิทธิ์ไฟล์เป็นแบบ '755' (rwxr-xr-x) ให้กับโฟลเดอร์เก็บเว็บ /var/www/html/ ควรใช้คำสั่งใด?",
        "options": [
            "sudo chmod 755 /var/www/html/",
            "sudo chown 755 /var/www/html/",
            "sudo chmod rwxr-xr-x /var/www/html/",
            "sudo chgrp 755 /var/www/html/"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "คำสั่งในข้อใดใช้เพื่อเปลี่ยนเจ้าของไฟล์ (Owner) ของ index.html ให้เป็นผู้ใช้ 'student01' และตั้งกลุ่มเป็น 'www-data'?",
        "options": [
            "sudo chmod student01:www-data index.html",
            "sudo chown student01:www-data index.html",
            "sudo setowner student01:www-data index.html",
            "sudo chown student01 www-data index.html"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "การใช้ SSH Key-based Authentication ในการล็อกอินเข้าเครื่องแม่ข่าย กุญแจส่วนใดที่ต้องติดตั้งไว้บนเครื่อง Server?",
        "options": [
            "Private Key (กุญแจส่วนตัว)",
            "Public Key (กุญแจสาธารณะ)",
            "Symmetric Key (กุญแจสมมาตร)",
            "Master Key (กุญแจหลัก)"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "การสร้างกุญแจคู่รหัสผ่าน SSH Key (Key Pair) บนเครื่องลูกข่าย (Client PC) นิยมใช้คำสั่งมาตรฐานใด?",
        "options": [
            "ssh-keygen -t ed25519",
            "ssh-copy-id -i key",
            "ssh-add -K private",
            "ssh-agent -s key"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "การทำ SSH Hardening ในไฟล์ /etc/ssh/sshd_config เพื่อห้ามใช้รหัสผ่านในการเข้าสู่ระบบทุกกรณี ต้องกำหนดตัวแปรใด?",
        "options": [
            "PasswordAuthentication no",
            "PermitRootLogin no",
            "AllowUsers keyonly",
            "DisablePassword Login yes"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "ข้อใดเป็นกฎเหล็กที่สำคัญที่สุดเมื่อคุณทำการแก้ไขไฟล์ตั้งค่า SSH และสั่ง reload บริการ SSH Server ใหม่?",
        "options": [
            "สั่งปิดการทำ Port Forwarding ของเราเตอร์ทันที",
            "ห้ามรันคำสั่ง sshd -t ก่อนการเริ่มระบบใหม่เด็ดขาด",
            "ห้ามปิด Terminal หน้าต่างเดิมจนกว่าจะทดสอบเปิดล็อกอินหน้าใหม่ผ่านสำเร็จ",
            "ลบไฟล์กุญแจ Private Key ทิ้งเพื่อเพิ่มความปลอดภัย"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "หากต้องการคัดลอกไฟล์ Public Key ของเครื่องเราไปยังเซิร์ฟเวอร์เพื่อให้ล็อกอินด้วย Key สำเร็จโดยอัตโนมัติ ควรใช้คำสั่งใด?",
        "options": [
            "ssh-copy-id",
            "ssh-add-key",
            "scp-key",
            "ssh-install"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "สิทธิ์ตัวเลขฐานแปด (Octal Mode) ของไฟล์ข้อความ '-rwx------' คือเลขใด?",
        "options": [
            "700",
            "755",
            "644",
            "600"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "หากหน้าเว็บไฟล์ index.html ในโฟลเดอร์ /var/www/html/ ถูกกำหนดสิทธิ์เป็น '600' (rw-------) และเป็นของ root ผู้ใช้งานทั่วไปที่เปิดเว็บเบราว์เซอร์จะพบข้อผิดพลาดใด?",
        "options": [
            "403 Forbidden (เนื่องจากผู้ใช้งาน Nginx/Others อ่านไฟล์ไม่ได้)",
            "404 Not Found (เนื่องจากไฟล์ถูกลบ)",
            "500 Internal Server Error (เนื่องจากสคริปต์พัง)",
            "200 OK (แสดงผลปกติ)"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "สิทธิ์แบบใดที่เรียกว่า 'Dangerous Permissions' (สิทธิ์อันตรายที่สุด) ที่ไม่ควรนำมาใช้บน Production Server โดยเด็ดขาด?",
        "options": [
            "777 (rwxrwxrwx)",
            "755 (rwxr-xr-x)",
            "644 (rw-r--r--)",
            "600 (rw-------)"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "คำสั่งใดใช้สลับบัญชีผู้ใช้งานจาก root ไปเป็นบัญชี 'student01' โดยดึงค่าสภาพแวดล้อมระบบ (Environment Variables) ของ student01 มาด้วย?",
        "options": [
            "su - student01",
            "su student01",
            "switch student01",
            "login student01"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "คำสั่งตรวจสอบความถูกต้องของไวยากรณ์ (Syntax Test) ในไฟล์ตั้งค่า SSH Server ก่อนสั่ง reload บริการคือคำสั่งใด?",
        "options": [
            "sudo sshd -t",
            "sudo systemctl check ssh",
            "sudo ssh -test",
            "sudo service sshd verify"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "หากต้องการโหลดไฟล์ตั้งค่าของ SSH Server ใหม่ (Reload) เพื่อให้การตั้งค่ามีผลโดยไม่มีการตัดการเชื่อมต่อของผู้ใช้ปัจจุบัน ควรใช้คำสั่งใด?",
        "options": [
            "sudo systemctl reload ssh",
            "sudo systemctl restart ssh",
            "sudo systemctl stop ssh",
            "sudo systemctl enable ssh"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "ในสิทธิ์ระบบลินุกซ์ สัญลักษณ์ 'o' (Others) หมายถึงกลุ่มเป้าหมายในข้อใด?",
        "options": [
            "ผู้ใช้ทุกคนที่เหลืออยู่ในระบบนอกเหนือจาก Owner และ Group",
            "เจ้าของไฟล์ (Owner)",
            "สมาชิกในกลุ่ม (Group)",
            "สิทธิ์พิเศษระบบ (root)"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "หากต้องการเปลี่ยนกลุ่มเจ้าของไฟล์ (Group Ownership) เพียงอย่างเดียวของไฟล์ report.txt โดยไม่เปลี่ยนเจ้าของไฟล์ตัวจริง ควรใช้คำสั่งใด?",
        "options": [
            "sudo chgrp www-data report.txt",
            "sudo chown www-data report.txt",
            "sudo chmod g+w report.txt",
            "sudo usermod -aG www-data report.txt"
        ],
        "correct": 1,
        "time": 25
    },
    {
        "q": "ไฟล์ที่ใช้เก็บรายชื่อ Public Keys ของคอมพิวเตอร์ภายนอกที่ได้รับอนุญาตให้ล็อกอินเข้าเซิร์ฟเวอร์ของแต่ละผู้ใช้ในระบบ จะถูกบันทึกไว้ในพาธใดบนเซิร์ฟเวอร์?",
        "options": [
            "~/.ssh/authorized_keys",
            "~/.ssh/id_ed25519.pub",
            "/etc/ssh/ssh_host_key",
            "/var/log/auth.log"
        ],
        "correct": 1,
        "time": 30
    }
]

def generate_xlsx(questions, output_path):
    try:
        wb = openpyxl.load_workbook('./documents/QuizizzSampleSpreadsheetUpdated_v2.xlsx')
        ws = wb.active
        
        # Clear all cells from row 3 downwards (up to 100)
        for r in range(3, 100):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None
                
        # Populate questions
        for idx, q_data in enumerate(questions):
            row = idx + 3
            ws.cell(row=row, column=1).value = q_data["q"]
            ws.cell(row=row, column=2).value = "Multiple Choice"
            ws.cell(row=row, column=3).value = q_data["options"][0]
            ws.cell(row=row, column=4).value = q_data["options"][1]
            ws.cell(row=row, column=5).value = q_data["options"][2]
            ws.cell(row=row, column=6).value = q_data["options"][3]
            ws.cell(row=row, column=7).value = None
            ws.cell(row=row, column=8).value = q_data["correct"]
            ws.cell(row=row, column=9).value = q_data["time"]
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None
            
        wb.save(output_path)
        print(f"Successfully generated {output_path} with {len(questions)} questions")
    except Exception as e:
        print(f"Error generating spreadsheet: {e}")

if __name__ == '__main__':
    generate_xlsx(w6a_questions, './public/data/week-6a_wayground_import.xlsx')
