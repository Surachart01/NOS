import openpyxl

w7b_pre_questions = [
    {
        "q": "UFW ย่อมาจากคำเต็มว่าอะไร?",
        "options": [
            "Universal File Wrapper",
            "Uncomplicated Firewall",
            "User Friendly Website",
            "Unified Flow Window"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "กฎ UFW เริ่มต้น (Default Policy) หลังจากติดตั้งมักถูกกำหนดไว้อย่างไรเพื่อความปลอดภัยสูงสุด?",
        "options": [
            "อนุญาตสัญญาณเข้าทั้งหมด และบล็อกสัญญาณออกทั้งหมด",
            "บล็อกสัญญาณเข้าทั้งหมด และอนุญาตสัญญาณออกทั้งหมด",
            "บล็อกทั้งสัญญาณเข้าและออกทั้งหมด",
            "เปิดบริการทุกพอร์ตโดยไม่มีการตรวจสอบใดๆ"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "คำสั่งใดใช้เพื่อเปิดใช้งานระบบ UFW Firewall ให้ทำงาน?",
        "options": [
            "sudo ufw start",
            "sudo ufw run",
            "sudo ufw enable",
            "sudo ufw service-start"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "หากต้องการบล็อกทุกการเชื่อมต่อจากไอพี 192.168.1.50 ไม่ให้เข้าเซิร์ฟเวอร์ของเรา ควรเขียนคำสั่งอย่างไร?",
        "options": [
            "sudo ufw block 192.168.1.50",
            "sudo ufw deny from 192.168.1.50",
            "sudo ufw reject to 192.168.1.50",
            "sudo ufw drop 192.168.1.50"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ถ้าต้องการเปิดพอร์ต 80 (HTTP) เฉพาะโปรโตคอล TCP ควรพิมพ์คำสั่งอย่างไร?",
        "options": [
            "sudo ufw allow 80/tcp",
            "sudo ufw open port 80",
            "sudo ufw permit tcp 80",
            "sudo ufw add rule 80 tcp"
        ],
        "correct": 1,
        "time": 20
    },
    {
        "q": "แนวคิด IP Whitelisting ในความปลอดภัยเครือข่ายหมายถึงอะไร?",
        "options": [
            "การลบรายชื่อไอพีทั้งหมดออกจากระบบล็อกการสแกน",
            "การเปลี่ยนไอพีแอดเดรสของเครื่องเซิร์ฟเวอร์ให้เป็นสีขาว",
            "การระบุรายชื่อไอพีที่ปลอดภัยเพื่อให้สิทธิ์เข้าถึงระบบได้เป็นกรณีพิเศษ",
            "การบล็อกไอพีจากทุกประเทศยกเว้นประเทศไทย"
        ],
        "correct": 3,
        "time": 30
    },
    {
        "q": "เครื่องมือ Nmap ทำงานสแกนพอร์ตแบบ SYN Stealth Scan (สแกนแบบเงียบ) ด้วยเทคนิคใด?",
        "options": [
            "ส่งแพ็กเก็ต SYN และรอรับ SYN-ACK แต่ไม่ส่ง ACK กลับเพื่อจบ Handshake",
            "ส่งไฟล์เปล่าขนาด 1MB ไปยังเครื่องเป้าหมายเพื่อวัดเวลา",
            "ทำการ Ping ไปยังเครื่องเป้าหมายอย่างต่อเนื่องเป็นเวลา 5 นาที",
            "แสร้งทำเป็นรันโปรแกรม Nginx เพื่อหลอกลวงบราวเซอร์ปลายทาง"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "ไฟล์ใดในเซิร์ฟเวอร์ Ubuntu ที่ทำหน้าที่บันทึกประวัติเหตุการณ์สกัดกั้นข้อมูลของ UFW (Firewall Log)?",
        "options": [
            "/var/log/nginx/access.log",
            "/var/log/syslog/ufw.log",
            "/var/log/ufw.log",
            "/etc/ufw/rules.log"
        ],
        "correct": 3,
        "time": 25
    },
    {
        "q": "หากต้องการดูบันทึก Log UFW แบบอัปเดตสดๆ ตลอดเวลาในขณะที่เพื่อนกำลังสแกนพอร์ตเรา ควรใช้คำสั่งใด?",
        "options": [
            "cat /var/log/ufw.log",
            "sudo tail -f /var/log/ufw.log",
            "sudo watch ufw.log",
            "less /var/log/ufw.log"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "คำสั่งใดใช้เพื่อตรวจสอบหมายเลขข้อ (Index) ของกฎ UFW เพื่อใช้อ้างอิงลบกฎเฉพาะข้อ?",
        "options": [
            "sudo ufw list --index",
            "sudo ufw status numbered",
            "sudo ufw show rules",
            "sudo ufw status verbose"
        ],
        "correct": 2,
        "time": 25
    }
]

w7b_post_questions = [
    {
        "q": "คำสั่งในข้อใดเป็นการอนุญาตเฉพาะเครื่องไอพี 192.168.1.150 เข้าถึงพอร์ตฐานข้อมูล 3306 (TCP) ได้ถูกต้องที่สุด?",
        "options": [
            "sudo ufw allow 3306/tcp to 192.168.1.150",
            "sudo ufw allow from 192.168.1.150 to any port 3306 proto tcp",
            "sudo ufw permit from 192.168.1.150 port 3306",
            "sudo ufw whitelist 192.168.1.150 on port 3306"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "เมื่อใช้คำสั่ง sudo ufw status แล้วได้ข้อความ Status: inactive หมายความว่าอย่างไร?",
        "options": [
            "ระบบ UFW กำลังค้นหาการ์ดเชื่อมต่อเครือข่ายใหม่",
            "UFW Firewall ทำงานปกติแต่ไม่พบบันทึกการเชื่อมต่อใหม่",
            "UFW Firewall ถูกปิดใช้งานอยู่และไม่มีผลบังคับใช้ใดๆ กับเครือข่าย",
            "ระบบปิดพอร์ตทุกพอร์ตยกเว้นพอร์ต 22 เพื่อความปลอดภัย"
        ],
        "correct": 3,
        "time": 25
    },
    {
        "q": "เหตุใดการลบกฎไฟร์วอลล์ด้วยลำดับตัวเลข (เช่น sudo ufw delete 3) จึงเป็นที่นิยมและปลอดภัยกว่าการลบด้วยคำสั่งกฎเต็ม?",
        "options": [
            "เพราะช่วยลดความเสี่ยงในการสะกดไวยากรณ์กฎผิดพลาดและลบได้รวดเร็ว",
            "เพื่อหลีกเลี่ยงไม่ให้ CPU ทำงานหนักเกินไปจากการค้นหาคำสั่ง",
            "เพื่อป้องกันไม่ให้แอปพลิเคชัน Nginx ตรวจเจอการเปลี่ยนแปลงระบบ",
            "เป็นกฎบังคับทางกฎหมายคอมพิวเตอร์ของการจัดการไฟร์วอลล์"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "คำสั่งใดใช้ในการปิดการทำงานของ UFW Firewall บนตู้ Ubuntu เซิร์ฟเวอร์อย่างถาวร?",
        "options": [
            "sudo ufw stop",
            "sudo ufw disable",
            "sudo ufw shutdown",
            "sudo ufw deactivate"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "หากเราสั่ง telnet [IP ปลายทาง] 80 แล้วขึ้นสถานะ Connection timed out ยาวนานโดยไม่มีการตอบกลับ หมายความว่าอะไร?",
        "options": [
            "พอร์ต 80 ของเซิร์ฟเวอร์ปิดการใช้งานและปฏิเสธสัญญาณทันที (REJECT)",
            "มีไฟร์วอลล์ของเซิร์ฟเวอร์เป้าหมายสกัดกั้นสัญญาณไว้ (DROP)",
            "เซิร์ฟเวอร์เป้าหมายกำลังรัน MariaDB ที่พอร์ต 80 แทน Nginx",
            "สายแลนเครือข่ายของเซิร์ฟเวอร์เป้าหมายถูกถอดออกทันที"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ใน UFW Log ไฟล์ /var/log/ufw.log ค่าตัวย่อ DPT และ SRC ย่อมาจากอะไรตามลำดับ?",
        "options": [
            "Data Packet Type และ Server Rule Code",
            "Database Port และ Security Router Connection",
            "Destination Port (พอร์ตปลายทาง) และ Source IP (ไอพีผู้ส่งต้นทาง)",
            "Direction Policy และ Source Port"
        ],
        "correct": 3,
        "time": 25
    },
    {
        "q": "หากต้องการจำลองสแกนพอร์ตเป้าหมายอย่างรวดเร็ว เฉพาะ 100 พอร์ตยอดนิยม ด้วย Nmap แบบ SYN Stealth Scan ควรพิมพ์คำสั่งอย่างไร?",
        "options": [
            "nmap -sS -F [IP]",
            "nmap -sV -p 1-65535 [IP]",
            "nmap --fast-scan [IP]",
            "nmap localhost -F"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "การบล็อก IP เดี่ยว (deny from 192.168.10.15) เทียบกับการบล็อก Subnet /24 (deny from 192.168.10.0/24) มีผลลัพธ์ต่างกันอย่างไร?",
        "options": [
            "การบล็อกแบบ Subnet จะสแกนหาเฉพาะเครื่องคอมพิวเตอร์ที่เป็นเซิร์ฟเวอร์หลัก",
            "การบล็อกแบบ Subnet จะปิดกั้นไอพีทั้งหมดตั้งแต่ .0 ถึง .255 ในวงแลนนั้น",
            "การบล็อกแบบเดี่ยวมีความปลอดภัยและทำงานได้เร็วกว่าการบล็อกแบบ Subnet 10 เท่า",
            "ไม่มีความแตกต่างกันในแง่ของจำนวนไอพีที่ถูกปิดกั้น"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "เพื่อป้องกันภัยคุกคามประเภท Brute-Force โจมตีพอร์ต 22 (SSH) ด้วยการสุ่มรหัสผ่าน UFW มีฟีเจอร์จำกัดความถี่การเชื่อมต่อที่เรียกว่าอะไร?",
        "options": [
            "ufw deny 22/tcp",
            "ufw limit 22/tcp",
            "ufw block-abuse 22",
            "ufw ssh-secure"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "กฎข้อใดของ UFW ในคลาสเรียนที่จะทำให้เพื่อนร่วมห้อง (ไอพี 192.168.1.180) ไม่สามารถสแกนพอร์ตหรือเชื่อมต่อใดๆ กับเครื่องคุณได้เลย?",
        "options": [
            "sudo ufw deny to 192.168.1.180",
            "sudo ufw deny from 192.168.1.180 to any",
            "sudo ufw block ip 192.168.1.180",
            "sudo ufw reject port all to 192.168.1.180"
        ],
        "correct": 2,
        "time": 30
    }
]

def generate_xlsx(questions, output_path):
    try:
        wb = openpyxl.load_workbook('./documents/QuizizzSampleSpreadsheetUpdated_v2.xlsx')
        ws = wb.active
        
        # Clear rows
        for r in range(3, 100):
            for c in range(1, 15):
                ws.cell(row=r, column=c).value = None
                
        # Populate
        for idx, q_data in enumerate(questions):
            row = idx + 3
            ws.cell(row=row, column=1).value = q_data["q"]
            ws.cell(row=row, column=2).value = "Multiple Choice"
            ws.cell(row=row, column=3).value = q_data["options"][0]
            ws.cell(row=row, column=4).value = q_data["options"][1]
            ws.cell(row=row, column=5).value = q_data["options"][2]
            ws.cell(row=row, column=6).value = q_data["options"][3]
            ws.cell(row=row, column=7).value = None
            ws.cell(row=row, column=8).value = q_data["correct"]
            ws.cell(row=row, column=9).value = q_data["time"]
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None
            
        wb.save(output_path)
        print(f"Successfully generated {output_path} with {len(questions)} questions")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == '__main__':
    generate_xlsx(w7b_pre_questions, './public/data/week-7b_pre_wayground.xlsx')
    generate_xlsx(w7b_post_questions, './public/data/week-7b_post_wayground.xlsx')
