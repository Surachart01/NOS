import openpyxl

w4a_questions = [
    {
        "q": "เทคโนโลยี Virtualization ในระบบเครือข่ายคอมพิวเตอร์มีจุดประสงค์หลักเพื่อทำสิ่งใด?",
        "options": [
            "ป้องกันระบบจากการบุกรุกของแฮกเกอร์",
            "จำลองแบ่งฮาร์ดแวร์เพื่อรันเครื่องคอมพิวเตอร์เสมือนหลายเครื่องพร้อมกัน",
            "เพิ่มความเร็วการเชื่อมต่ออินเทอร์เน็ตของเครื่องลูกข่าย",
            "บีบอัดขนาดไฟล์ข้อมูลระบบทั้งหมดให้มีขนาดเล็กลง"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "Hypervisor ประเภทใดที่ถูกติดตั้งลงบนฮาร์ดแวร์ของเครื่องเซิร์ฟเวอร์โดยตรง (Bare-metal) โดยไม่ต้องรันภายใต้ระบบปฏิบัติการอื่น?",
        "options": [
            "Type 1 Hypervisor",
            "Type 2 Hypervisor",
            "Type 3 Hypervisor",
            "Hosted Hypervisor"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "ข้อใดจัดเป็นตัวอย่างของ Type 1 (Bare-metal) Hypervisor ที่นิยมใช้เป็นระบบจัดการเซิร์ฟเวอร์เสมือนในองค์กร?",
        "options": [
            "VirtualBox",
            "VMware Workstation",
            "Proxmox VE",
            "Parallels Desktop"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "ซอฟต์แวร์จำลองเครื่องเสมือนประเภทใดที่มีความเบาหวิว (Lightweight) และแบ่งปันการใช้งานระบบ Kernel ร่วมกับระบบหลัก?",
        "options": [
            "KVM (Kernel-based Virtual Machine)",
            "LXC (Linux Containers)",
            "Windows Server VM",
            "Hyper-V VM"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "โดเมนของหน้าจอควบคุมระบบ Web UI ของ Proxmox VE ทำงานอยู่ภายใต้พอร์ตหลักหมายเลขใดและเชื่อมด้วยโปรโตคอลใด?",
        "options": [
            "HTTP พอร์ต 80",
            "HTTPS พอร์ต 443",
            "HTTPS พอร์ต 8006",
            "HTTP พอร์ต 8006"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "หากต้องการติดตั้งเซิร์ฟเวอร์จริงขึ้นมาเป็น Proxmox VE สิ่งใดเป็นเงื่อนไขสำคัญที่ต้องเข้าไปเปิดใช้งานในหน้าจอ BIOS ก่อนทำการลงระบบ?",
        "options": [
            "เปิดใช้ระบบ DHCP Server",
            "เปิดใช้งานเทคโนโลยีระบบเสมือน (VT-x หรือ AMD-V)",
            "เปิดใช้สิทธิ์การ์ดอินเทอร์เฟส Wi-Fi",
            "เปิดการทำ RAID 0 เสมอ"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ในหน้าจอติดตั้งระบบช่วง Network Configuration ตัวแปร IP Address ที่เรากรอกจะต้องกำหนดในรูปแบบใดจึงจะถูกต้องตามความต้องการของ Proxmox?",
        "options": [
            "กำหนด IP แบบสุ่มโดยใช้ DHCP",
            "กำหนด IP แบบคงที่ (Static IP) พร้อมรหัส CIDR เช่น 192.168.10.50/24",
            "กรอกเฉพาะไอพีโดยไม่ต้องใส่ Mask เช่น 192.168.10.50",
            "ใช้หมายเลข IP เดียวกันกับเครื่องเราเตอร์หลัก"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "หลังการบูตติดตั้ง Proxmox VE สำเร็จและเครื่องทำการ Reboot ตัวเอง สิ่งแรกที่คุณควรดำเนินการคือข้อใด?",
        "options": [
            "กดสวิตช์ปิดหน้าจอทันทีเพื่อรอสัญญาณเชื่อมต่อ",
            "ถอดแฟลชไดรฟ์ USB บูตออกเพื่อไม่ให้ระบบวนกลับไปหน้าต่างติดตั้งใหม่",
            "สั่งปิดพอร์ต 8006 บนเราเตอร์ของสถาบัน",
            "พิมพ์รหัสผ่าน root ซ้ำๆ บนหน้าจอบูตเพื่อเข้ารหัส"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "ข้อความเตือนความปลอดภัย \"Not Secure\" หรือ Certificate Error เมื่อเปิดเบราว์เซอร์เข้าลิงก์ Proxmox เกิดขึ้นจากสาเหตุใด?",
        "options": [
            "ระบบเซิร์ฟเวอร์โดนแฮกข้อมูลและไม่ปลอดภัย",
            "ระบบใช้ใบรับรองความปลอดภัยแบบลงนามเอง (Self-signed Certificate) ซึ่งเป็นเรื่องปกติของเครื่องภายในแลน",
            "สัญญาณสายแลนของเครื่องนักเรียนไม่แน่น",
            "รหัสผ่านผู้ใช้งาน root มีการพิมพ์ผิดพลาด"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ระบบปฏิบัติการหลัก (Base OS) ของตัวควบคุม Proxmox VE มีฐานข้อมูลซอฟต์แวร์และการรันคำสั่งทำงานอยู่บน Linux ค่ายใด?",
        "options": [
            "CentOS",
            "Alpine Linux",
            "Red Hat Enterprise",
            "Debian"
        ],
        "correct": 4,
        "time": 20
    }
]

w4b_questions = [
    {
        "q": "หน้าที่หลักของซอฟต์แวร์ Web Server เช่น Nginx ในระบบเครือข่ายคือข้อใด?",
        "options": [
            "เชื่อมต่อสายไฟอินเทอร์เน็ตเข้ามายังอาคาร",
            "รับส่งแพ็กเกจ HTTP Request และคืนผลลัพธ์เป็นหน้าเว็บเพจ (HTTP Response)",
            "แจกจ่ายหมายเลข IP Address ให้แก่เครื่องลูกข่ายโดยอัตโนมัติ",
            "ล้างหน้าจอระบบเพื่อลบประวัติการควบคุมระบบทั้งหมด"
        ],
        "correct": 2,
        "time": 30
    },
    {
        "q": "ช่องทางพอร์ตมาตรฐานในการรับส่งข้อมูลบริการเว็บไซต์ธรรมดา (HTTP) และเว็บแบบเข้ารหัสปลอดภัย (HTTPS) คือพอร์ตหมายเลขใดตามลำดับ?",
        "options": [
            "พอร์ต 22 และ พอร์ต 80",
            "พอร์ต 80 และ พอร์ต 443",
            "พอร์ต 443 และ พอร์ต 8006",
            "พอร์ต 80 และ พอร์ต 8080"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "โปรโตคอล SSH (Secure Shell) ซึ่งใช้ควบคุมเซิร์ฟเวอร์ระยะไกลแบบปลอดภัยทางคอนโซล ทำงานอยู่บนช่องพอร์ตมาตรฐานใด?",
        "options": [
            "พอร์ต 80",
            "พอร์ต 443",
            "พอร์ต 22",
            "พอร์ต 8006"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "หากต้องการเชื่อมต่อ SSH เข้าสู่เครื่องเซิร์ฟเวอร์ตู้เสมือนของตนเองด้วยสิทธิ์ root บนไอพี 192.168.10.150 ควรป้อนคำสั่งอย่างไร?",
        "options": [
            "ssh 192.168.10.150 -root",
            "ssh root@192.168.10.150",
            "connect ssh root to 192.168.10.150",
            "ssh root:192.168.10.150"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "ก่อนทำคำสั่งติดตั้ง Nginx บนระบบลินุกซ์ด้วยคำสั่ง sudo apt install nginx ข้อปฏิบัติข้อใดที่ระบบควรทำก่อนเสมอ?",
        "options": [
            "สั่งปิดพอร์ต 22 เพื่อรักษาความปลอดภัย",
            "รันคำสั่ง sudo apt update เพื่ออัปเดตรายการดัชนีแอปพลิเคชันล่าสุด",
            "สั่งลบโฟลเดอร์ผู้ใช้อื่นออกจากไดเรกทอรี /home",
            "ทำการเปลี่ยนชื่อ Hostname เป็น nginx.local"
        ],
        "correct": 2,
        "time": 25
    },
    {
        "q": "คำสั่งใดที่ระบบใช้เพื่อเปิดให้ Nginx เริ่มต้นทำงานใหม่อีกครั้งทันทีแบบล้างโปรเซสทำงานเก่าหลังแก้ไขตัวแปร?",
        "options": [
            "sudo systemctl status nginx",
            "sudo systemctl stop nginx",
            "sudo systemctl restart nginx",
            "sudo systemctl enable nginx"
        ],
        "correct": 3,
        "time": 20
    },
    {
        "q": "พิกัดโฟลเดอร์หรือไดเรกทอรีมาตรฐาน (Default Document Root) ที่ Nginx ใช้เก็บไฟล์หน้าเว็บ index.html คือพิกัดใด?",
        "options": [
            "/etc/nginx/html/",
            "/var/www/html/",
            "/home/student/html/",
            "/usr/share/nginx/"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "เหตุใดการเขียนหรือลบไฟล์ index.html ในโฟลเดอร์ /var/www/html/ จึงจำเป็นต้องใช้คำสั่งนำหน้าด้วย sudo?",
        "options": [
            "เพราะโฟลเดอร์นี้ถูกรันด้วยสิทธิ์ความปลอดภัยสูงสุดและจำกัดสิทธิ์เฉพาะ root เท่านั้น",
            "เพื่อทำให้หน้าเว็บเปิดแสดงผลกราฟิกสีสันได้รวดเร็วยิ่งขึ้น",
            "เพื่อหลีกเลี่ยงข้อจำกัดการเชื่อมต่อพอร์ต 8006",
            "เป็นกฎบังคับของการใช้บราวเซอร์ Google Chrome เสมอ"
        ],
        "correct": 1,
        "time": 30
    },
    {
        "q": "กฎเหล็กของแอดมินลินุกซ์ ในการตรวจสอบความถูกต้องของไวยากรณ์ในไฟล์ตั้งค่าของ Nginx ก่อนสั่งเริ่มระบบใหม่คือคำสั่งใด?",
        "options": [
            "sudo systemctl test nginx",
            "sudo nginx -t",
            "check -nginx config",
            "verify nginx.conf"
        ],
        "correct": 2,
        "time": 20
    },
    {
        "q": "หากสั่งแก้ไขไฟล์คอนฟิกของ Nginx ผิดไวยากรณ์ไป 1 ตัวอักษร แล้วรันคำสั่ง restart บริการ ผลกระทบในข้อใดจะเกิดขึ้นกับเซิร์ฟเวอร์?",
        "options": [
            "ระบบจะข้ามข้อความที่สะกดผิดไปรันค่าเดิมอัตโนมัติโดยไม่มีผลกระทบ",
            "บริการ Nginx จะล้มเหลวในการเปิดตัวและหยุดการทำงานลงทันที (เว็บล่ม)",
            "การ์ดเครือข่ายของเซิร์ฟเวอร์จะปิดตัวและยกเลิกรับค่า IP Address",
            "หน้าจอเทอร์มินัลจะสั่งลบไฟล์เก็บข้อมูลเว็บทิ้งถาวร"
        ],
        "correct": 2,
        "time": 30
    }
]

def generate_xlsx(questions, output_path):
    wb = openpyxl.load_workbook('./documents/QuizizzSampleSpreadsheetUpdated_v2.xlsx')
    ws = wb.active
    
    # Clear all cells from row 3 downwards (up to 100)
    for r in range(3, 100):
        for c in range(1, 15):
            ws.cell(row=r, column=c).value = None
            
    # Populate questions
    for idx, q_data in enumerate(questions):
        row = idx + 3
        ws.cell(row=row, column=1).value = q_data["q"]
        ws.cell(row=row, column=2).value = "Multiple Choice"
        ws.cell(row=row, column=3).value = q_data["options"][0]
        ws.cell(row=row, column=4).value = q_data["options"][1]
        ws.cell(row=row, column=5).value = q_data["options"][2]
        ws.cell(row=row, column=6).value = q_data["options"][3]
        ws.cell(row=row, column=7).value = None
        ws.cell(row=row, column=8).value = q_data["correct"]
        ws.cell(row=row, column=9).value = q_data["time"]
        ws.cell(row=row, column=10).value = None
        ws.cell(row=row, column=11).value = None
        
    wb.save(output_path)
    print(f"Successfully generated {output_path}")

generate_xlsx(w4a_questions, './public/data/week-4a_wayground_import.xlsx')
generate_xlsx(w4b_questions, './public/data/week-4b_wayground_import.xlsx')
